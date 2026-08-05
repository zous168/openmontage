"""FAQ 导入模板本机落盘."""

from __future__ import annotations

from pathlib import Path

import pytest

from plugins.mxai.kb.faq_template_local import (
    TEMPLATE_FILENAME,
    build_faq_import_template_xlsx,
    ensure_faq_template_file,
    list_faq_templates,
)


def test_build_template_has_headers() -> None:
    raw = build_faq_import_template_xlsx()
    assert len(raw) > 100
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(raw), read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0][:2] == ("question", "answer")
    assert rows[1][0] and rows[1][1]


def test_ensure_and_list(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("USERPROFILE", str(tmp_path))
    # Path.home() 在部分环境不读 USERPROFILE；直接改 faq_template_dir
    monkeypatch.setattr(
        "plugins.mxai.kb.faq_template_local.faq_template_dir",
        lambda: tmp_path / "Documents" / "MxAI" / "FAQ模板",
    )
    item = ensure_faq_template_file()
    assert item["name"] == TEMPLATE_FILENAME
    assert Path(item["path"]).is_file()
    listed = list_faq_templates(ensure=True)
    assert listed["dir"]
    assert any(i["name"] == TEMPLATE_FILENAME for i in listed["items"])
    # 不覆盖已有
    Path(item["path"]).write_bytes(b"keep")
    again = ensure_faq_template_file(overwrite=False)
    assert Path(again["path"]).read_bytes() == b"keep"
