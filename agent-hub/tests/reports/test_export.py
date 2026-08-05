"""LT-002.08.1：报表生成与导出."""

from io import BytesIO

from openpyxl import load_workbook


def test_report_generate_export(mxai_client) -> None:
    gen = mxai_client.post(
        "/api/plugins/mxai/reports/generate",
        json={"report_type": "daily"},
    )
    assert gen.status_code == 200
    report_id = gen.json()["report_id"]
    detail = mxai_client.get(f"/api/plugins/mxai/reports/{report_id}")
    assert detail.status_code == 200
    export = mxai_client.post(f"/api/plugins/mxai/reports/{report_id}/export")
    assert export.status_code == 200
    assert export.content[:2] == b"PK"
    wb = load_workbook(BytesIO(export.content), read_only=True)
    assert wb.sheetnames == ["汇总", "原始明细"]
    ws_sum = wb["汇总"]
    assert ws_sum["A1"].value == "报表ID"
    assert ws_sum["B1"].value == report_id
    ws_det = wb["原始明细"]
    headers = [cell.value for cell in next(ws_det.iter_rows(min_row=1, max_row=1))]
    assert headers[0] == "明细ID"
    assert "操作时间" in headers
